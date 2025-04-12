from pathlib import Path
import importlib
import pandas as pd
import json
import os
from sklearn.metrics import classification_report
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side, BORDER_THICK, BORDER_THIN
import matplotlib
import shutil
from sklearn.metrics import f1_score


class TaskEvalResSummarizer:
    def __init__(self, task_name:str, method:str=None, labels:list=None) -> None:
        '''
        '''
        self.labels = labels
        self.task_name = task_name
        self.method = method
        print(f'Initial method value: {method}')  # Debug print
        print(f'logic: {method is None}')  # Debug print
        if method is None: # summarize all method groups
            task_dir = Path(f"./results/{task_name}")
            print('!!!!!!!!!!!!!!Here1')
            for method_dir in task_dir.iterdir():
                if method_dir.is_dir():
                    self.method_dir = method_dir
                    self.summarize_eval_res()
        else:
            print('!!!!!!!!!!!!!!Here2')
            self.method_dir = Path(f"./results/{task_name}/{method}")
            self.summarize_eval_res()

    def create_eval_cols(self):
        col_names = [
            'model_name',  'test_set', 'none_nr', 'acc', 'marco_avg_f1', 'weighted_avg_f1', 'precision', 'recall', 'f1'
                    #'nomulti_acc', 'nomulti_marco_avg_f1', 'nomulti_weighted_avg_f1',
                    #'multi_acc', 'multi_marco_avg_f1', 'multi_weighted_avg_f1',
                    #'sub_acc', 'sub_marco_avg_f1','sub_weighted_avg_f1',
                    #'sub2_acc', 'sub2_marco_avg_f1','sub2_weighted_avg_f1'
                    ]
        for label in self.labels:
            col_names.extend([f'{label}_p', f'{label}_r', f'{label}_f1', f'{label}_support'])
        return col_names
    
    def group_models_by_name(self):
        model_names = [d.name for d in self.method_dir.iterdir() if d.is_dir()]
        # baselines = [d for d in model_names if '_lora' not in d]
        # model_names = [d for d in model_names if d not in baselines]

        # models_7B = [d for d in model_names if '7B_' in d]
        # models_7B_Chat = [d for d in model_names if '7B-Chat_' in d]
        # models_13B = [d for d in model_names if '13B_' in d]
        # models_13B_Chat = [d for d in model_names if '13B-Chat_' in d]
        # models_8b = [d for d in model_names if '8b_' in d]
        # models_8b_instr = [d for d in model_names if '8b-Instruct' in d]
        # baselines.sort()
        # models_7B.sort()
        # models_7B_Chat.sort()
        # models_13B.sort()
        # models_13B_Chat.sort()

        # model_names = models_13B_Chat + models_13B + models_7B_Chat + models_7B + baselines + models_8b + models_8b_instr

        return model_names
    
    def get_eval_dict(self, model_dir, model_name, col_names):
        dic = {}
        for col in col_names:
            dic[col] = ''
                
        eval_file = model_dir/ 'eval_report.json'
        pred_file = model_dir/ 'eval_pred.csv'
        cm_file = model_dir/ 'eval_cm.csv'
        pred_df = pd.read_csv(pred_file)
        with open(eval_file, 'r') as f:
            eval_dict = json.load(f)
        cm_df = pd.read_csv(cm_file)
        replacement_dict = {'no': 0, 'yes': 1}
        pred_df['true'] = pred_df['true'].apply(lambda x: replacement_dict.get(x, 0))  # Default to 'no' => 0
        pred_df['pred'] = pred_df['pred'].apply(lambda x: replacement_dict.get(x, 0))  # Default to 'no' => 0
        f1_sl = f1_score(pred_df['true'], pred_df['pred'], average='binary')
        TN = cm_df.iloc[0, 1]
        FP = cm_df.iloc[0, 2]
        FN = cm_df.iloc[1, 1]
        TP = cm_df.iloc[1, 2]
        precision = TP / (TP + FP)
        recall = TP / (TP + FN)
        F1 = 2 * (precision * recall) / (precision + recall)        
        if 'accuracy' not in eval_dict.keys():
            acc = 0
        else:
            acc = round(eval_dict['accuracy'], 4)
        dic['model_name'] = model_name
        dic['test_set'] = len(pred_df)
        dic['none_nr'] = eval_dict['none_nr']
        dic['acc'] = acc
        dic['marco_avg_f1'] = round(eval_dict['macro avg']['f1-score'], 4)
        dic['weighted_avg_f1'] = round(eval_dict['weighted avg']['f1-score'], 4)
        dic['precision'] = precision
        dic['recall'] = recall
        dic['f1'] = f1_sl
        # dic['f1_score_sklearn'] = f1_sl

        for label in self.labels:
            dic[f'{label}_p'] = round(eval_dict[label]['precision'], 4)
            dic[f'{label}_r'] = round(eval_dict[label]['recall'], 4)
            dic[f'{label}_f1'] = round(eval_dict[label]['f1-score'], 4)
            dic[f'{label}_support'] = eval_dict[label]['support']

        return dic
    
    def save_df_to_excel(self, excelfile, sheetname, df):
        print(df[['model_name', 'test_set', 'acc', 'marco_avg_f1', 'weighted_avg_f1', 'precision','recall','f1']])
        
        if not os.path.isfile(excelfile):
            #print('The excel file is not existing, creating a new excel file...', excelfile)
            wb = Workbook()
            wb.save(excelfile)
            excelfile_work = excelfile
        else:
            new_excelfile = excelfile.parent/f'{excelfile.stem}_new.xlsx'
            shutil.copy(excelfile, new_excelfile)
            excelfile_work = new_excelfile

        '''
        excelfile = shutil.copy(excelfile, excelfile)
        writer = pd.ExcelWriter(
                excelfile,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="replace",
            )
        

        wb = load_workbook(excelfile)
        if not (sheetname in wb.sheetnames):
            #print('The worksheet is not existing, creating a new worksheet...' , sheetname)
            ws1 = wb.create_sheet(sheetname)
            ws1.title = sheetname
            wb.save(excelfile)
        '''
        writer = pd.ExcelWriter(
                excelfile_work,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="replace",
            )

        #book = load_workbook(excelfile)
        #idx = wb.sheetnames.index(sheetname)
        #ws = book.get_sheet_by_name(sheetname)
        #book.remove(ws)
        #book.create_sheet(sheetname, idx)
        #writer = pd.ExcelWriter(excelfile, engine='openpyxl')
        #writer.book = book
        #writer.sheets = {ws.title: ws for ws in book.worksheets}

        df.to_excel(writer, sheet_name=sheetname, index=False, header=True)
        writer.close()
        #writer.save()

        if excelfile_work != excelfile:
            shutil.copy(excelfile_work, excelfile)
            os.remove(excelfile_work)

    def check_best_models(self, df, metrics):
        best_models = {}
        for m in metrics:
            best_models.update({m: []})

        for m in metrics:
            # remove ''
            df = df[df[m]!='']
            max_v = df[m].max()
            for index, row in df.iterrows():
                if row[m] == max_v:
                    best_models[m].append((row['model_name'], row['test_set']))
        return best_models
    
    
    
    def color_best_metric(self, excelfile, sheetname, best_models, color, font, metrics):
        wb = load_workbook(excelfile)
        ws = wb[sheetname]

        # Create a dictionary of column names
        ColNames = {}
        Current = 0
        for COL in ws.iter_cols(1, ws.max_column):
            ColNames[COL[0].value] = Current
            Current += 1

        # Color best metrics
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for m in metrics:
                bests = best_models[m]
                for model in bests:
                    if row_cells[ColNames['model_name']].value == model[0] and row_cells[ColNames['test_set']].value == model[1]:
                            row_cells[ColNames[m]].fill = PatternFill("solid", fgColor=color)
                            row_cells[ColNames[m]].font = Font(b=font)
                    if not 'ckp' in row_cells[ColNames['model_name']].value:
                        row_cells[ColNames['model_name']].fill = PatternFill("solid", fgColor='FFCCFF')

        wb.save(excelfile)

    def summarize_eval_res(self):
        res_dir = self.method_dir
        print('summarizing the results....', res_dir)
        eval_res_dir = Path(f'./eval_results/{self.task_name}')
        eval_res_dir.mkdir(parents=True, exist_ok=True)
        eval_res_dir = eval_res_dir/self.method_dir.name
        eval_res_dir.mkdir(parents=True, exist_ok=True)
        eval_summary_file = eval_res_dir/f'{res_dir.name}_eval_summ.xlsx'
        col_names = self.create_eval_cols()
        eval_summary = pd.DataFrame(columns=col_names)
        
        #group models by name
        model_names = self.group_models_by_name()

        for model_name in model_names:
                model_dir = res_dir/model_name
                model_ckp_dirs = [d for d in model_dir.iterdir() if d.is_dir() and 'checkpoint' in d.name]
                # sort by the number of the checkpoint
                model_ckp_dirs.sort(key=lambda x: int(x.name.split('-')[-1]))
                eval_file = model_dir/ 'eval_report.json'
                eval_file_cm = model_dir/ 'eval_cm.csv'
                print('### summarizing: ', model_dir)
                if eval_file.exists() and eval_file_cm.exists():
                    dic = self.get_eval_dict(model_dir, model_name, col_names)
                    df = pd.DataFrame(dic, index=[0])                  
                    eval_summary = pd.concat([eval_summary, df], ignore_index=True)
                if len(model_ckp_dirs) > 0:
                    for ckp_dir in model_ckp_dirs:
                        ckp_nr = ckp_dir.name.split('-')[-1]
                        eval_file = ckp_dir/ 'eval_report.json'
                        if not eval_file.exists():
                            print('--- ckp not evaluated', ckp_dir.name)
                            continue
                        dic = self.get_eval_dict(ckp_dir, f'{model_name}_ckp{ckp_nr}', col_names)
                        df = pd.DataFrame(dic, index=[0])
                        eval_file_cm = ckp_dir/ 'eval_cm.csv'
                        eval_summary = pd.concat([eval_summary, df], ignore_index=True)
        all_df = eval_summary
        baseline_df = eval_summary[~eval_summary['model_name'].str.contains('_lora')]
        self.save_df_to_excel(eval_summary_file, 'all', all_df)
        self.save_df_to_excel(eval_summary_file, 'all_baselines', baseline_df)

        green = '00FF00'
        lightgreen = 'CCFFCC'
        lightblue = '00FFFF'
        pink = 'FF00FF'
        yellow = 'FFFF00'
        lightyellow = 'FFFFCC'
        red = 'FF0000'
        metric_cols = col_names[3:]
        metric_cols = [c for c in metric_cols if 'support' not in c]
        all_bests = self.check_best_models(all_df, metric_cols)
        self.color_best_metric(eval_summary_file, 'all', all_bests, green, True, metric_cols)
        baseline_bests = self.check_best_models(baseline_df, metric_cols)
        self.color_best_metric(eval_summary_file, 'all_baselines', baseline_bests, lightblue, True, metric_cols)
        
        # make a copy of the summary file, rename it with the current time
        new_file = eval_summary_file.parent/f'{eval_summary_file.name}_{datetime.now().strftime("%Y%m%d-%H%M%S")}.xlsx'
        shutil.copy(eval_summary_file, new_file)
        # remove the old summary file
        os.remove(eval_summary_file)
        
        
                                             

    

def main():
    ''

if __name__ == "__main__":
    main()